# jira_relatorio.py
from __future__ import annotations

import io
import os
import re
import smtplib
import logging
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email import encoders
from datetime import datetime, timedelta
from typing import List, Dict, Any

import pandas as pd
import requests
from dateutil import tz
from flask import Flask, jsonify, request
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger

# -----------------------------------------------------------------------------
# Configuração
# -----------------------------------------------------------------------------
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(name)s - %(message)s")
logger = logging.getLogger("jira_relatorio")

TZ = tz.gettz(os.getenv("APP_TZ", "America/Sao_Paulo"))

SMTP_HOST = os.getenv("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER")
SMTP_PASS = os.getenv("SMTP_PASS")
EMAIL_FROM = os.getenv("EMAIL_FROM", SMTP_USER or "noreply@example.com")
EMAIL_TO_DEFAULT = [e.strip() for e in os.getenv("EMAIL_TO", "").split(",") if e.strip()]

JIRA_BASE_URL = os.getenv("JIRA_BASE_URL")
JIRA_USER = os.getenv("JIRA_USER")
JIRA_TOKEN = os.getenv("JIRA_TOKEN")

# Projeto fixo
PROJECT_KEY = "IDT"

# JQL: status traduzido "Concluído" OU categoria Done
JIRA_JQL_TEMPLATE = (
    'project = IDT '
    'AND (status = "Concluído" OR statusCategory = Done) '
    'AND (resolved >= -{days}d OR status changed to "Concluído" after -{days}d) '
    'ORDER BY resolved DESC, updated DESC'
)

# Custom fields
CF_SETOR = "customfield_10156"
CF_FINALIZADO = "customfield_10009"
CF_FILIAL = "customfield_10147"

app = Flask(__name__)
scheduler = BackgroundScheduler(timezone=TZ)
scheduler.start()

# -----------------------------------------------------------------------------
# Helpers
# -----------------------------------------------------------------------------
def _normalize_cf(value):
    """Normaliza valores de custom fields do Jira (dict/list/str)."""
    if value is None:
        return None
    if isinstance(value, dict):
        for k in ("displayName", "name", "value", "title"):
            if k in value and value[k]:
                return str(value[k])
        return str(value)
    if isinstance(value, list):
        vals = [_normalize_cf(v) for v in value]
        return ", ".join([v for v in vals if v])
    return str(value)

def _strip_tz_inplace(df: pd.DataFrame, local_tz):
    """Remove timezone de colunas datetime (Excel exige naive)."""
    for col in df.columns:
        if pd.api.types.is_datetime64tz_dtype(df[col]):
            df[col] = df[col].dt.tz_convert(local_tz).dt.tz_localize(None)
    for col in df.columns:
        if df[col].dtype == "object":
            sample = df[col].dropna().head(5)
            if sample.empty:
                continue

            def _normalize(x):
                if isinstance(x, pd.Timestamp):
                    return (x.tz_convert(local_tz).tz_localize(None) if x.tz is not None else x).to_pydatetime()
                try:
                    ts = pd.to_datetime(x, errors="raise", utc=True)
                    return ts.tz_convert(local_tz).tz_localize(None).to_pydatetime()
                except Exception:
                    return x

            df[col] = df[col].apply(_normalize)

def _sanitize_sheet_name(name: str) -> str:
    """Planilha Excel: máx 31 chars e sem : \ / ? * [ ]"""
    name = name or "Sem Setor"
    name = re.sub(r'[:\\/\?\*\[\]]', "-", name)
    return name[:31]

# -----------------------------------------------------------------------------
# Colunas e cabeçalhos (Criado em ANTES de Finalizado)
# -----------------------------------------------------------------------------
COLUMN_ORDER = [
    "key", "summary", "status", "assignee", "reporter",
    "setor", "filial", "created", "finalizado"
]
HEADER_PTBR = {
    "key": "ID",
    "summary": "Resumo",
    "status": "Status",
    "assignee": "Responsável",
    "reporter": "Autor",
    "setor": "Setor",
    "filial": "Filial",
    "created": "Criado em",
    "finalizado": "Finalizado",
}

# -----------------------------------------------------------------------------
# Estilo + AutoFit de colunas
# -----------------------------------------------------------------------------
def _apply_excel_style(ws, df_cols_ptbr: List[str]):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Cabeçalho
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for c, _ in enumerate(df_cols_ptbr, 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Bordas finas
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border

    # Datas: dd/mm/yyyy hh:mm
    for col_name in ["Criado em", "Finalizado"]:
        if col_name in df_cols_ptbr:
            idx = df_cols_ptbr.index(col_name) + 1
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=idx).number_format = "dd/mm/yyyy hh:mm"

    # Filtros e congelar topo
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # -------- AutoFit de colunas (largura conforme conteúdo) --------
    # Largura mínima por coluna (em "nº de caracteres") e teto global
    PREFERRED_MIN = {
        "ID": 12,
        "Resumo": 60,
        "Status": 12,
        "Responsável": 22,
        "Autor": 22,
        "Setor": 24,
        "Filial": 28,
        "Criado em": 19,     # "dd/mm/yyyy hh:mm" ~ 16 chars
        "Finalizado": 19,
    }
    MIN_W = 8
    MAX_W = 120  # teto bem alto pra não cortar textos grandes

    for col_idx, header in enumerate(df_cols_ptbr, start=1):
        column_letter = get_column_letter(col_idx)
        max_len = len(str(header))  # inclui o próprio cabeçalho
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=col_idx).value
            if val is None:
                continue
            if isinstance(val, datetime):
                length = 16
            else:
                # considera quebras de linha; pega o trecho mais comprido
                parts = str(val).splitlines()
                length = max(len(p) for p in parts) if parts else 0
            if length > max_len:
                max_len = length

        base = max(PREFERRED_MIN.get(header, MIN_W), max_len + 2)
        ws.column_dimensions[column_letter].width = min(MAX_W, base)
    # ----------------------------------------------------------------

# -----------------------------------------------------------------------------
# Jira
# -----------------------------------------------------------------------------
def fetch_tasks_from_jira(jql: str, fields: List[str] | None = None, max_results: int = 2000) -> List[Dict[str, Any]]:
    if not (JIRA_BASE_URL and JIRA_USER and JIRA_TOKEN):
        logger.warning("Jira não configurado; retornando lista vazia.")
        return []

    url = f"{JIRA_BASE_URL.rstrip('/')}/rest/api/3/search"
    headers = {"Accept": "application/json"}
    auth = (JIRA_USER, JIRA_TOKEN)

    if fields is None:
        fields = [
            "key", "summary", "status", "assignee", "reporter",
            "created", "updated", "resolutiondate",
            "issuetype", "priority", "project",
            CF_SETOR, CF_FINALIZADO, CF_FILIAL,
        ]

    start_at, page_size, items = 0, 100, []
    while start_at < max_results:
        payload = {"jql": jql, "startAt": start_at, "maxResults": min(page_size, max_results - start_at), "fields": fields}
        try:
            resp = requests.post(url, json=payload, headers=headers, auth=auth, timeout=60)
            resp.raise_for_status()
        except requests.HTTPError as e:
            status = getattr(e.response, "status_code", None)
            try:
                err_body = e.response.json()
            except Exception:
                err_body = getattr(e.response, "text", str(e))
            logger.error("Jira search %s. JQL=%s | payload=%s | error=%s", status, jql, payload, err_body)
            raise

        data = resp.json()
        for issue in data.get("issues", []):
            f = issue.get("fields", {})
            items.append({
                "key": issue.get("key"),
                "summary": f.get("summary"),
                "status": (f.get("status") or {}).get("name") if f.get("status") else None,
                "assignee": (f.get("assignee") or {}).get("displayName") if f.get("assignee") else None,
                "reporter": (f.get("reporter") or {}).get("displayName") if f.get("reporter") else None,
                "created": f.get("created"),
                "updated": f.get("updated"),
                "resolved": f.get("resolutiondate"),
                "setor": _normalize_cf(f.get(CF_SETOR)),
                "finalizado": _normalize_cf(f.get(CF_FINALIZADO)),
                "filial": _normalize_cf(f.get(CF_FILIAL)),
            })

        total = data.get("total", 0)
        start_at += page_size
        if start_at >= total or start_at >= max_results:
            break

    logger.info("Jira retornou %d itens", len(items))
    return items

# -----------------------------------------------------------------------------
# DataFrame
# -----------------------------------------------------------------------------
def build_dataframe(items: List[Dict[str, Any]]) -> pd.DataFrame:
    df = pd.DataFrame(items)
    if df.empty:
        return pd.DataFrame(columns=COLUMN_ORDER)

    # Converte datas (UTC -> TZ)
    for col in ["created", "updated", "resolved", "finalizado"]:
        if col in df.columns:
            s = pd.to_datetime(df[col], errors="coerce", utc=True)
            if s.notna().any():
                s = s.dt.tz_convert(TZ)
            df[col] = s

    # Remove timezone para Excel
    _strip_tz_inplace(df, TZ)

    # Reordena e ordena por Finalizado (desc) ou Criado em (desc)
    cols_present = [c for c in COLUMN_ORDER if c in df.columns]
    df = df[cols_present]
    if "finalizado" in df.columns and df["finalizado"].notna().any():
        df = df.sort_values(by=["finalizado"], ascending=False)
    elif "created" in df.columns:
        df = df.sort_values(by=["created"], ascending=False)
    return df

# -----------------------------------------------------------------------------
# Excel (abas: Todos os Setores + por Setor)
# -----------------------------------------------------------------------------
def dataframe_to_excel_bytes_grouped_by_setor(df: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Aba consolidada
        df_all = (df.copy() if not df.empty else pd.DataFrame(columns=COLUMN_ORDER))
        df_all_export = df_all.rename(columns=HEADER_PTBR)
        df_all_export.to_excel(writer, index=False, sheet_name="Todos os Setores")
        _apply_excel_style(writer.sheets["Todos os Setores"], list(df_all_export.columns))

        # Abas por setor
        if df.empty:
            empty = pd.DataFrame(columns=COLUMN_ORDER).rename(columns=HEADER_PTBR)
            empty.to_excel(writer, index=False, sheet_name="Setor - Sem Setor")
            _apply_excel_style(writer.sheets["Setor - Sem Setor"], list(empty.columns))
        else:
            for setor_val, gdf in df.groupby("setor", dropna=False):
                nome_setor = setor_val if pd.notna(setor_val) and str(setor_val).strip() else "Sem Setor"
                sheet = _sanitize_sheet_name(f"Setor - {nome_setor}")
                df_export = gdf.copy().rename(columns=HEADER_PTBR)
                df_export.to_excel(writer, index=False, sheet_name=sheet)
                _apply_excel_style(writer.sheets[sheet], list(df_export.columns))
    return output.getvalue()

# -----------------------------------------------------------------------------
# E-mail
# -----------------------------------------------------------------------------
def send_mail_with_attachment(subject: str, body_text: str, to_emails: List[str], attachment_name: str, attachment_bytes: bytes):
    if not to_emails:
        raise ValueError("Lista de destinatários vazia.")
    if not (SMTP_HOST and SMTP_PORT and SMTP_USER and SMTP_PASS):
        raise RuntimeError("SMTP não configurado corretamente (host/port/user/pass).")

    msg = MIMEMultipart()
    msg["From"], msg["To"], msg["Subject"] = EMAIL_FROM, ", ".join(to_emails), subject
    msg.attach(MIMEText(body_text, "plain", "utf-8"))

    part = MIMEBase("application", "octet-stream")
    part.set_payload(attachment_bytes)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f"attachment; filename={attachment_name}")
    msg.attach(part)

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
        server.starttls()
        server.login(SMTP_USER, SMTP_PASS)
        server.sendmail(EMAIL_FROM, to_emails, msg.as_string())

    logger.info("E-mail enviado para: %s", to_emails)

# -----------------------------------------------------------------------------
# Orquestração
# -----------------------------------------------------------------------------
def run_report_and_email(days: int = 30, jql: str | None = None, to_emails: List[str] | None = None, dry: bool = False) -> Dict[str, Any]:
    to_emails = to_emails or EMAIL_TO_DEFAULT

    jql_final = (jql or JIRA_JQL_TEMPLATE.format(days=days))
    items = fetch_tasks_from_jira(jql_final) if jql_final else []

    # Filtro defensivo por janela (preferindo Finalizado -> Resolved -> Updated -> Created)
    if items:
        cutoff = datetime.now(tz=TZ) - timedelta(days=days)
        keep = []
        for it in items:
            dt_finalizado = pd.to_datetime(it.get("finalizado"), errors="coerce", utc=True)
            dt_resolved   = pd.to_datetime(it.get("resolved"),   errors="coerce", utc=True)
            dt_updated    = pd.to_datetime(it.get("updated"),    errors="coerce", utc=True)
            dt_created    = pd.to_datetime(it.get("created"),    errors="coerce", utc=True)
            chosen = next((d for d in (dt_finalizado, dt_resolved, dt_updated, dt_created) if pd.notna(d)), None)
            if chosen is not None and chosen.tzinfo is not None:
                chosen = chosen.tz_convert(TZ)
            if chosen is not None and chosen >= cutoff:
                keep.append(it)
        items = keep

    df = build_dataframe(items)
    excel_bytes = dataframe_to_excel_bytes_grouped_by_setor(df)

    now = datetime.now(tz=TZ)
    filename = f"relatorio_{PROJECT_KEY.lower()}_{now.strftime('%Y%m%d_%H%M')}.xlsx"

    if dry:
        return {"ok": True, "dry": True, "rows": int(len(df.index)), "filename": filename}

    subject = f"Relatório {PROJECT_KEY} - tarefas concluídas (últimos {days} dias) - {now.strftime('%d/%m/%Y %H:%M')}"
    body = (
        f"Segue em anexo o relatório em Excel do projeto {PROJECT_KEY}.\n"
        f"Gerado em {now.strftime('%d/%m/%Y %H:%M %Z')}."
    )
    send_mail_with_attachment(subject, body, to_emails, filename, excel_bytes)

    return {"ok": True, "rows": int(len(df.index)), "filename": filename, "sent_to": to_emails, "generated_at": now.isoformat()}

# -----------------------------------------------------------------------------
# API
# -----------------------------------------------------------------------------
@app.get("/health")
def health():
    return jsonify({"status": "ok", "time": datetime.now(tz=TZ).isoformat(), "project": PROJECT_KEY})

@app.post("/report/run")
def report_run():
    data = request.get_json(silent=True) or {}
    days = int(data.get("days", 30))
    jql = data.get("jql")  # override opcional
    to_emails = data.get("to")
    if isinstance(to_emails, str):
        to_emails = [e.strip() for e in to_emails.split(",") if e.strip()]
    dry = str(data.get("dry", "")).lower() in {"1", "true", "yes"}

    try:
        result = run_report_and_email(days=days, jql=jql, to_emails=to_emails, dry=dry)
        return jsonify(result)
    except Exception as e:
        logger.exception("Erro executando relatório: %s", e)
        return jsonify({"ok": False, "error": str(e)}), 500

# ---- agendamentos (opcionais) ------------------------------------------------
@app.post("/schedule/monthly")
def schedule_monthly():
    data = request.get_json(force=True)
    day = int(data.get("day", 1))
    time_str = data.get("time", "09:00")
    days_window = int(data.get("days", 30))
    jql = data.get("jql")
    to_emails = data.get("to") or EMAIL_TO_DEFAULT
    if isinstance(to_emails, str):
        to_emails = [e.strip() for e in to_emails.split(",") if e.strip()]
    h, m = map(int, time_str.split(":"))
    job_id = data.get("job_id", f"monthly_{day:02d}_{h:02d}{m:02d}")
    trigger = CronTrigger(day=day, hour=h, minute=m, timezone=TZ)

    def job():
        logger.info("[JOB %s] Executando relatório mensal...", job_id)
        run_report_and_email(days=days_window, jql=jql, to_emails=to_emails)

    if scheduler.get_job(job_id):
        scheduler.remove_job(job_id)
    scheduler.add_job(job, trigger=trigger, id=job_id, replace_existing=True)
    return jsonify({"ok": True, "job_id": job_id, "when": str(trigger)})

@app.post("/schedule/daily")
def schedule_daily():
    data = request.get_json(force=True)
    time_str = data.get("time", "09:00")
    days_window = int(data.get("days", 30))
    jql = data.get("jql")
    to_emails = data.get("to") or EMAIL_TO_DEFAULT
    if isinstance(to_emails, str):
        to_emails = [e.strip() for e in to_emails.split(",") if e.strip()]
    h, m = map(int, time_str.split(":"))
    job_id = data.get("job_id", f"daily_{h:02d}{m:02d}")
    trigger = CronTrigger(hour=h, minute=m, timezone=TZ)

    def job():
        logger.info("[JOB %s] Executando relatório diário...", job_id)
        run_report_and_email(days=days_window, jql=jql, to_emails=to_emails)

    if scheduler.get_job(job_id):
        scheduler.remove_job(job_id)
    scheduler.add_job(job, trigger=trigger, id=job_id, replace_existing=True)
    return jsonify({"ok": True, "job_id": job_id, "when": str(trigger)})

@app.get("/schedule")
def schedule_list():
    jobs = []
    for job in scheduler.get_jobs():
        jobs.append({
            "id": job.id,
            "next_run_time": job.next_run_time.astimezone(TZ).isoformat() if job.next_run_time else None,
            "trigger": str(job.trigger),
        })
    return jsonify({"ok": True, "jobs": jobs})

@app.delete("/schedule/<job_id>")
def schedule_delete(job_id: str):
    job = scheduler.get_job(job_id)
    if not job:
        return jsonify({"ok": False, "error": "job_id não encontrado"}), 404
    scheduler.remove_job(job_id)
    return jsonify({"ok": True, "removed": job_id})

# -----------------------------------------------------------------------------
if __name__ == "__main__":
    port = int(os.getenv("PORT", "8000"))
    logger.info("Iniciando API na porta %d (TZ=%s, PROJECT=%s)...", port, TZ, PROJECT_KEY)
    app.run(host="0.0.0.0", port=port)

